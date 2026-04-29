# ============================================================
# step2_excel.py
# WHAT  : Save weather data into an Excel file
# HOW   : Use openpyxl to create/append rows to a .xlsx file
# ============================================================

# ============================================================
# step2_excel.py
# WHAT  : Save forex rates data into an Excel file
# HOW   : Use openpyxl to create/append rows to a .xlsx file
# ============================================================
import openpyxl
import os
from datetime import datetime

OUTPUT_FOLDER = "output"
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "forex_report.xlsx")

HEADERS = [
    "ID",
    "Amount",
    "Base Currency",
    "Target Currency",
    "Exchange Rate",
    "API Date",
    "Scraped At"
]

def save_to_excel(data: dict) -> str:
    """
    Save forex rate data into output/forex_report.xlsx.
    Creates folder/file if not exist. Adds data row for each target currency.
    Returns the file path.
    """
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Open or create workbook
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        next_id = ws.max_row  # because row 1 is headers
        print(f"[EXCEL] Appending to existing file. Current rows: {ws.max_row - 1}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Forex Report"
        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        next_id = 2
        print("[EXCEL] Created new Excel file with headers.")

    # Add data rows for each rate
        for target, rate in data["rates"].items():
            row = [
                next_id,
                1.0,
                data["base"],
                target,
                float(rate),
                data["date"],
                data.get("scraped_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]
            ws.append(row)
            next_id += 1

    wb.save(EXCEL_FILE)
    print(f"[EXCEL] Saved rates → {EXCEL_FILE}")
    return EXCEL_FILE

# Test run directly
if __name__ == "__main__":
    from step1_scrape import scrape_country   # Your step 1 script
    forex_data = scrape_country()
    if forex_data:
        save_to_excel(forex_data)









































# import openpyxl
# import os
# from datetime import datetime


# # Output folder and file name
# OUTPUT_FOLDER = "output"
# EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "weather_report.xlsx")

# # Column headers for the Excel sheet
# HEADERS = [
#     "ID",
#     "City",
#     "Country",
#     "Temperature (°C)",
#     "Feels Like (°C)",
#     "Humidity (%)",
#     "Wind Speed (km/h)",
#     "Weather Description",
#     "Visibility (km)",
#     "Scraped At",
# ]


# def save_to_excel(weather_data: dict) -> str:
#     """
#     Append weather data to an Excel file.
#     Creates the file with headers if it doesn't exist yet.
#     Returns the file path.
#     """

#     # Make sure the output folder exists
#     os.makedirs(OUTPUT_FOLDER, exist_ok=True)

#     # ── Load existing workbook OR create a new one ────────────
#     if os.path.exists(EXCEL_FILE):
#         wb = openpyxl.load_workbook(EXCEL_FILE)
#         ws = wb.active
#         next_id = ws.max_row  # existing rows + 1 (header counts as row 1)
#         print(f"[EXCEL] Appending to existing file. Current rows: {ws.max_row - 1}")
#     else:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Weather Report"

#         # ── Style the header row ──────────────────────────────
#         from openpyxl.styles import Font, PatternFill, Alignment

#         header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
#         header_font = Font(color="FFFFFF", bold=True)

#         ws.append(HEADERS)

#         for col_num, cell in enumerate(ws[1], 1):
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = Alignment(horizontal="center")

#         # Set column widths
#         col_widths = [6, 16, 16, 18, 16, 14, 18, 22, 15, 22]
#         for i, width in enumerate(col_widths, 1):
#             ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

#         next_id = 1
#         print("[EXCEL] Created new Excel file with headers.")

#     # ── Append the new data row ───────────────────────────────
#     row = [
#         next_id,
#         weather_data["city"],
#         weather_data["country"],
#         weather_data["temperature_c"],
#         weather_data["feels_like_c"],
#         weather_data["humidity_percent"],
#         weather_data["wind_speed_kmph"],
#         weather_data["weather_desc"],
#         weather_data["visibility_km"],
#         weather_data["scraped_at"],
#     ]
#     ws.append(row)

#     wb.save(EXCEL_FILE)
#     print(f"[EXCEL] Saved row ID {next_id} → {EXCEL_FILE}")
#     return EXCEL_FILE


# # ── Run this file directly to test ───────────────────────────
# if __name__ == "__main__":
#     # Import step 1 scraper for testing
#     from step1_scrape import scrape_weather

#     data = scrape_weather("KualaLumpur")
#     file_path = save_to_excel(data)
#     print(f"\n[EXCEL] File saved at: {file_path}")